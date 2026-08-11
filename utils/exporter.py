"""ReportExporter — 报表导出工具

依赖:
    openpyxl  — Excel 导出
    matplotlib — 图表生成

两个公开方法均接受 teaching_id + file_path，
由 View 层负责弹出文件保存对话框后调用。
"""
from __future__ import annotations

from datetime import datetime
from typing import Optional

from core.exceptions import BusinessError
from models.analysis_dao import AnalysisDAO
from services.grade_service import GradeService


class ReportExporter:
    def __init__(self) -> None:
        self._grade_svc = GradeService()
        self._analysis_dao = AnalysisDAO()

    # ------------------------------------------------------------------
    # Excel 成绩单
    # ------------------------------------------------------------------
    def export_class_grade_sheet(self, teaching_id: int, file_path: str) -> None:
        """将指定授课班级的成绩单导出为格式化 Excel 文件。

        Args:
            teaching_id: edu_teaching.id
            file_path:   保存路径（.xlsx）

        Raises:
            BusinessError: 无成绩数据或写文件失败。
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import (
                Alignment, Border, Font, PatternFill, Side,
            )
        except ImportError as exc:
            raise BusinessError("请先安装 openpyxl: pip install openpyxl") from exc

        grades = self._grade_svc.get_teaching_grades(teaching_id)
        stats = self._analysis_dao.get_teaching_stats(teaching_id)

        wb = Workbook()
        ws = wb.active
        ws.title = "成绩单"

        # 标题行
        course_name = stats.get("course_name", "") if stats else ""
        semester = stats.get("semester", "") if stats else ""
        ws.merge_cells("A1:G1")
        title_cell = ws["A1"]
        title_cell.value = f"{course_name}  {semester} 成绩单"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")

        # 列标题
        headers = ["序号", "学号", "姓名", "成绩", "等级", "绩点", "是否通过"]
        header_fill = PatternFill("solid", fgColor="4472C4")
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        # 数据行
        for idx, row in enumerate(grades, 1):
            data = [
                idx,
                row.get("username", ""),
                row.get("real_name", ""),
                float(row["score"]) if row.get("score") is not None else "—",
                row.get("grade_letter") or "—",
                float(row["gpa_point"]) if row.get("gpa_point") is not None else "—",
                "通过" if row.get("is_passed") else "不通过",
            ]
            row_fill = PatternFill("solid", fgColor="DCE6F1") if idx % 2 == 0 else None
            for col, val in enumerate(data, 1):
                cell = ws.cell(row=idx + 2, column=col, value=val)
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
                if row_fill:
                    cell.fill = row_fill

        # 统计摘要行
        if stats:
            summary_row = len(grades) + 4
            ws.cell(row=summary_row, column=1, value="班级统计").font = Font(bold=True)
            summary_data = [
                ("平均分", stats.get("avg_score")),
                ("最高分", stats.get("max_score")),
                ("最低分", stats.get("min_score")),
                ("标准差", stats.get("std_dev")),
                ("及格率", f"{stats.get('pass_rate', 0)}%"),
            ]
            for offset, (label, val) in enumerate(summary_data):
                ws.cell(row=summary_row + 1, column=offset + 1, value=label).font = Font(bold=True)
                ws.cell(row=summary_row + 2, column=offset + 1, value=val)

        # 列宽自适应
        col_widths = [6, 12, 10, 8, 8, 8, 10]
        for col, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

        try:
            wb.save(file_path)
        except OSError as exc:
            raise BusinessError(f"文件保存失败: {exc}") from exc

    # ------------------------------------------------------------------
    # 分析报告（PDF / PNG）
    # ------------------------------------------------------------------
    def export_analysis_report(
        self,
        teaching_id: int,
        file_path: str,
        title: Optional[str] = None,
    ) -> None:
        """生成成绩分析报告（含分段柱状图 + 正态曲线），保存为 PDF 或 PNG。

        Args:
            teaching_id: edu_teaching.id
            file_path:   保存路径（.pdf 或 .png）
            title:       图表标题（可选）

        Raises:
            BusinessError: 无成绩数据或保存失败。
        """
        try:
            import matplotlib
            matplotlib.use("Agg")  # 非交互式后端，确保线程安全
            import matplotlib.pyplot as plt
            import numpy as np
            from matplotlib import rcParams
            rcParams["font.sans-serif"] = ["Microsoft YaHei", "SimHei", "DejaVu Sans"]
            rcParams["axes.unicode_minus"] = False
        except ImportError as exc:
            raise BusinessError("请先安装 matplotlib: pip install matplotlib") from exc

        dist = self._analysis_dao.get_score_distribution(teaching_id)
        if not dist or not dist.get("total"):
            raise BusinessError("该班级暂无成绩数据，无法生成报告。")

        stats = self._analysis_dao.get_teaching_stats(teaching_id)
        course_name = stats.get("course_name", "") if stats else ""
        semester = stats.get("semester", "") if stats else ""
        chart_title = title or f"{course_name} {semester} 成绩分析报告"

        # 分段数据
        segments = ["<60", "60-69", "70-79", "80-89", "90-100"]
        counts = [
            int(dist.get("below_60", 0) or 0),
            int(dist.get("cnt_60_69", 0) or 0),
            int(dist.get("cnt_70_79", 0) or 0),
            int(dist.get("cnt_80_89", 0) or 0),
            int(dist.get("cnt_90_100", 0) or 0),
        ]
        colors = ["#E74C3C", "#E67E22", "#F1C40F", "#2ECC71", "#3498DB"]

        fig, axes = plt.subplots(1, 2, figsize=(12, 5))
        fig.suptitle(chart_title, fontsize=14, fontweight="bold")

        # --- 左图：分段柱状图 ---
        ax1 = axes[0]
        bars = ax1.bar(segments, counts, color=colors, edgecolor="white", linewidth=0.8)
        ax1.set_title("成绩分段分布", fontsize=12)
        ax1.set_xlabel("分数段")
        ax1.set_ylabel("人数")
        ax1.set_ylim(0, max(counts) * 1.3 + 1)
        for bar, cnt in zip(bars, counts):
            if cnt > 0:
                ax1.text(
                    bar.get_x() + bar.get_width() / 2,
                    bar.get_height() + 0.1,
                    str(cnt),
                    ha="center",
                    va="bottom",
                    fontweight="bold",
                )

        # 统计信息文本框
        avg = float(dist.get("avg_score") or 0)
        std = float(dist.get("std_dev") or 0)
        pass_rate = float(dist.get("pass_rate") or 0)
        info_text = (
            f"均分: {avg:.1f}\n"
            f"最高: {dist.get('max_score', '—')}\n"
            f"最低: {dist.get('min_score', '—')}\n"
            f"标准差: {std:.2f}\n"
            f"及格率: {pass_rate:.1f}%"
        )
        ax1.text(
            0.98, 0.97, info_text,
            transform=ax1.transAxes,
            verticalalignment="top",
            horizontalalignment="right",
            bbox={"boxstyle": "round", "facecolor": "lightyellow", "alpha": 0.8},
            fontsize=9,
        )

        # --- 右图：正态分布拟合曲线 ---
        ax2 = axes[1]
        if avg > 0 and std > 0:
            x = np.linspace(max(0, avg - 4 * std), min(100, avg + 4 * std), 200)
            y = np.exp(-0.5 * ((x - avg) / std) ** 2) / (std * np.sqrt(2 * np.pi))
            y_scaled = y * int(dist.get("total", 1)) * 10  # 缩放到人数量级
            ax2.plot(x, y_scaled, color="#3498DB", linewidth=2, label="正态拟合")
            ax2.fill_between(x, y_scaled, alpha=0.15, color="#3498DB")
            ax2.axvline(avg, color="#E74C3C", linestyle="--", linewidth=1.2, label=f"均分 {avg:.1f}")
            ax2.set_title("成绩正态分布拟合", fontsize=12)
            ax2.set_xlabel("分数")
            ax2.set_ylabel("频率（估计）")
            ax2.legend(fontsize=9)
        else:
            ax2.text(0.5, 0.5, "数据不足，无法拟合", ha="center", va="center",
                     transform=ax2.transAxes, fontsize=12, color="gray")
            ax2.set_title("成绩正态分布拟合", fontsize=12)

        plt.tight_layout()

        try:
            plt.savefig(file_path, dpi=150, bbox_inches="tight")
        except OSError as exc:
            raise BusinessError(f"文件保存失败: {exc}") from exc
        finally:
            plt.close(fig)
